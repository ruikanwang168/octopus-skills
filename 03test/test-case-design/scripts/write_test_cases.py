#!/usr/bin/env python3
"""将标准化中文测试用例从 JSON 写入 XLSX 或 CSV。"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path
from typing import Any


HEADERS = [
    "所属产品",
    "所属模块",
    "相关研发需求",
    "用例名称",
    "前置条件",
    "关键词",
    "优先级",
    "用例类型",
    "适用阶段",
    "用例状态",
    "步骤",
    "预期",
]


class ChineseArgumentParser(argparse.ArgumentParser):
    def format_usage(self) -> str:
        return super().format_usage().replace("usage:", "用法:", 1)

    def format_help(self) -> str:
        return super().format_help().replace("usage:", "用法:", 1)


def normalize_text(value: str) -> str:
    text = value.strip()
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    return (
        text.replace("\r\n", "\n")
        .replace("\r", "\n")
        .replace("\\r\\n", "\n")
        .replace("\\n", "\n")
        .replace("\\r", "\n")
    )


def format_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return normalize_text(value)
    if isinstance(value, (int, float, bool)):
        return str(value)
    if isinstance(value, list):
        lines: list[str] = []
        for index, item in enumerate(value, start=1):
            if isinstance(item, dict):
                text = item.get("text") or item.get("内容") or item.get("步骤") or item.get("预期") or item
                lines.append(f"{index}. {format_cell(text)}")
            else:
                text = format_cell(item)
                lines.append(text if text.startswith(f"{index}.") else f"{index}. {text}")
        return "\n".join(lines)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return str(value)


def load_cases(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        data = data.get("cases") or data.get("用例") or data.get("test_cases")
    if not isinstance(data, list):
        raise ValueError("输入 JSON 必须是列表，或是包含 cases 数组的对象。")

    cases: list[dict[str, Any]] = []
    for index, item in enumerate(data, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"第 {index} 条用例必须是对象。")
        cases.append({header: format_cell(item.get(header, "")) for header in HEADERS})
    return cases


def write_csv(cases: list[dict[str, str]], output: Path) -> Path:
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(cases)
    return output


def write_xlsx(cases: list[dict[str, str]], output: Path) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        fallback = output.with_suffix(".csv")
        write_csv(cases, fallback)
        print(
            f"警告：未安装 openpyxl，已改为写入 CSV 文件：{fallback}",
            file=sys.stderr,
        )
        return fallback

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "测试用例"
    sheet.append(HEADERS)

    for case in cases:
        sheet.append([case[header] for header in HEADERS])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "所属产品": 16,
        "所属模块": 18,
        "相关研发需求": 28,
        "用例名称": 32,
        "前置条件": 30,
        "关键词": 18,
        "优先级": 10,
        "用例类型": 12,
        "适用阶段": 14,
        "用例状态": 12,
        "步骤": 44,
        "预期": 44,
    }
    for index, header in enumerate(HEADERS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = widths[header]

    for row in sheet.iter_rows(min_row=2):
        max_lines = 1
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, str):
                max_lines = max(max_lines, len(cell.value.splitlines()) or 1)
        sheet.row_dimensions[row[0].row].height = min(18 + (max_lines - 1) * 16, 180)

    sheet.freeze_panes = "A2"
    workbook.save(output)
    return output


def main() -> int:
    parser = ChineseArgumentParser(
        description="将测试用例 JSON 写入 XLSX 或 CSV。",
        add_help=False,
        usage="%(prog)s 输入文件 [--output 输出文件]",
    )
    parser._positionals.title = "位置参数"
    parser._optionals.title = "可选参数"
    parser.add_argument("input", type=Path, metavar="输入文件", help="包含生成测试用例的 JSON 文件。")
    parser.add_argument("-h", "--help", action="help", help="显示帮助信息并退出。")
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        metavar="输出文件",
        default=Path("测试用例.xlsx"),
        help="输出 .xlsx 或 .csv 路径，默认生成 测试用例.xlsx。",
    )
    args = parser.parse_args()

    try:
        cases = load_cases(args.input)
        suffix = args.output.suffix.lower()
        if suffix == ".csv":
            written = write_csv(cases, args.output)
        elif suffix == ".xlsx":
            written = write_xlsx(cases, args.output)
        else:
            raise ValueError("输出路径必须以 .xlsx 或 .csv 结尾。")
    except Exception as exc:
        print(f"错误：{exc}", file=sys.stderr)
        return 1

    print(f"已写入 {len(cases)} 条用例：{written}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
